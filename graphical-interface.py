import tkinter as tk
from tkinter import ttk
from tkinter import PhotoImage
from tkinter import filedialog
import tkinter.messagebox as messagebox
import os
from PIL import Image, ImageTk
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
from myStyle import myStyles
import serial
from serial.tools import list_ports
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from  openpyxl.chart import LineChart, Reference, ScatterChart, Series



class App:
 
    def __init__(self, root):

        myStyles()

        self.root = root   
        self.serial_obj = None
        self.stop = True
        self.enter = True
        self.zeropos = 0
        self.pos = tk.StringVar()
        self.load= tk.StringVar()
        self.arraypos = []
        self.arrayforce = []
        self.helpLblVar = tk.StringVar()
        self.default_help_message = "Collegare USB e selezionare porta COM per interagire col banco prova."
        self.helpLblVar.set(self.default_help_message)

        # Create the style object
        self.style = ttk.Style()  

        # Get the screen dimensions for adaptive sizing
        self.screen_width = root.winfo_screenwidth()
        self.screen_height = root.winfo_screenheight()
        window_width = int(self.screen_width * 0.8)
        window_height = int(self.screen_height * 0.8)

        # GUI setup with adaptive size
        root.title("Prova a trazione")
        root.geometry(f"{window_width}x{window_height}")

        # Configuring columns and rowsto expand dynamically
        columnsNum = 6
        rowNum = 7

        for i in range(columnsNum):
            root.grid_columnconfigure(i, weight=1)

        root.grid_rowconfigure(0, weight=1)
        root.grid_rowconfigure(1, weight=1)
        for i in range(2, rowNum):
            root.grid_rowconfigure(i, weight=2)
        

        """HEADER SECTION"""
        self.headerFrame = ttk.Frame(root, padding=20, style='headerFrame.TFrame', )
        self.headerFrame.grid(row=0, column=0, columnspan=columnsNum, sticky="nsew")

        # Make the header frame expand vertically and horizontally
        self.headerFrame.grid_rowconfigure(0, weight=1)
        self.headerFrame.grid_columnconfigure(0, weight=1)
        self.headerFrame.grid_columnconfigure(1, weight=4)
        self.headerFrame.grid_columnconfigure(2, weight=1)


        # Load the original image
        self.original_image = Image.open("img/logo_waya-removebg.png")


        # Get the original dimensions of the image
        original_width, original_height = self.original_image.size

        # Define the new width or height (adjust this as needed)
        self.new_height = int(self.screen_height / 20) # should be adaptive to the screen height this way
        self.aspect_ratio = original_height / original_width

        # Calculate the new height while maintaining the aspect ratio
        self.new_width = int(self.new_height / self.aspect_ratio)

        # Resize the image
        image_resized = self.original_image.resize((self.new_width, self.new_height))

        # Convert the image to a Tkinter-compatible format
        self.logo = ImageTk.PhotoImage(image_resized)

        self.logo_label = ttk.Label(self.headerFrame, image=self.logo, anchor="center", style="headerLabel.TLabel")
        self.logo_label.grid(column=0, row=0, sticky="nswe", padx=5, pady=5)

        # Bind window resize event
        self.root.bind("<Configure>", self.resize_logo)
     
        # ttk.Label(self.headerFrame, image=self.logo, anchor="center", style="headerLabel.TLabel").grid(column=0, row=0, sticky="nswe", padx=5, pady=5)
        ttk.Label(self.headerFrame, text="Interfaccia banco per prove rumorosità", anchor="center", font=("Arial", 16, "bold"), style="headerLabel.TLabel").grid(column=1, row=0, sticky="nswe", padx=5, pady=5)

        root.grid_rowconfigure(1, minsize=3, weight=0)
        # Create an orange bottom border
        self.headerBorder = ttk.Frame(root, style="headerBorder.TFrame")
        self.headerBorder.grid(column=0, row=1, columnspan=columnsNum, sticky="nsew")

        """PANED WINDOW FOR THE LOWER SECTION"""
        self.pw = ttk.PanedWindow(root, orient ='horizontal', style="custom.TPanedwindow")
        self.pw.grid(row=2, column=0, rowspan=rowNum-2, columnspan=columnsNum, sticky="nsew")


        """BUTTONS & LABELS SECTION"""
        self.buttonsFrame = ttk.Frame(self.pw, padding=20)
        self.pw.add(self.buttonsFrame, weight=2)

        # Make the buttons frame expand vertically and horizontally
        self.buttonsFrame.grid_rowconfigure(0, weight=1)
        self.buttonsFrame.grid_columnconfigure(0, weight=1)


        self.zerosFrame = ttk.Frame(self.buttonsFrame, style="MyCustomFrame.TFrame")
        self.zerosFrame.grid(row=0, column=0, sticky="nsew")

        for i in range(4):
            self.zerosFrame.grid_rowconfigure(i, weight=1)
        self.zerosFrame.grid_rowconfigure(4, weight=2)

        self.zerosFrame.grid_columnconfigure(0, weight=3)
        self.zerosFrame.grid_columnconfigure(1, weight=2)


        self.comLabel = ttk.Label(self.zerosFrame, text="Selezionare porta 'COM':", font=("Arial", 16), anchor='e', style="bodyLabel.TLabel").grid(row=0, column=0, sticky="ew")

        # Create Combobox and set available COM ports
        self.com_ports = []
        self.com_port_var = tk.StringVar()

        self.com_ports_dropdown = ttk.Combobox(self.zerosFrame, textvariable=self.com_port_var, values=self.com_ports, state="readonly", style="custom.TCombobox")
        self.com_ports_dropdown.grid(row=0, column=1, pady=5)        

        # Get available COM ports
        self.refresh_com_ports()

        # Set first available port as default (if any)
        if self.com_ports:
            self.com_port_var.set(self.com_ports[0])
        else:
            self.com_port_var.set("")


        # Associa la selezione della porta COM all'evento di connessione
        self.com_ports_dropdown.bind("<<ComboboxSelected>>", self.connect_serial)

        self.connect_serial()

        # Sposta la help label qui, sotto la combobox
        self.helpLbl = ttk.Label(self.zerosFrame, textvariable=self.helpLblVar, font=("Arial", 12), background="#F1F1F1", anchor="w", padding=(0,20))
        self.helpLbl.grid(row=1, column=0, columnspan=2, sticky='ew')

        # Aggiorna le righe successive
        next_row = 2

        self.startBtn = ttk.Button(self.zerosFrame, text='Start', style="startBtn.TButton", command=self.start_acquisition)
        self.startBtn.grid(row=next_row, column=0)

        self.stopBtn = ttk.Button(self.zerosFrame, text='Stop', style="stopBtn.TButton", command=self.stop_acquisition)
        self.stopBtn.grid(row=next_row, column=1)


        """GRAPH SECTION"""
        self.graphFrame = ttk.Frame(self.pw, style="MyCustomFrame.TFrame", padding=20)
        self.pw.add(self.graphFrame, weight=4)

        # Make the graph frame expand vertically and horizontally
        self.graphFrame.grid_columnconfigure(0, weight=1)
        self.graphFrame.grid_rowconfigure(0, weight=1)
        self.graphFrame.grid_rowconfigure(1, weight=4)

        # --- NUOVO: Frame superiore per posizione e forza, affiancati ---
        self.topGraphFrame = ttk.Frame(self.graphFrame, style="MyCustomFrame.TFrame")
        self.topGraphFrame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        self.topGraphFrame.grid_columnconfigure(0, weight=1)
        self.topGraphFrame.grid_columnconfigure(1, weight=1)
        self.topGraphFrame.grid_rowconfigure(0, weight=1)

        self.lblfrmPos = ttk.Labelframe(self.topGraphFrame, text='Posizione Attuale [mm]', style="lblFrm.TLabelframe")
        self.lblfrmPos.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.posLabel = ttk.Label(self.lblfrmPos, textvariable=self.pos, font=("Arial", 20), anchor="center", background='#F1F1F1')
        self.posLabel.pack(expand=True)

        self.lblfrmForce = ttk.Labelframe(self.topGraphFrame, text='Forza Attuale [N]', style="lblFrm.TLabelframe")
        self.lblfrmForce.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        self.forceLabel = ttk.Label(self.lblfrmForce, textvariable=self.load, font=("Arial", 20), anchor="center", background='#F1F1F1')
        self.forceLabel.pack(expand=True)
        # --- FINE NUOVO ---

        # Add buttons
        self.buttonSaveFrame = ttk.Frame(self.graphFrame, style="MyCustomFrame.TFrame", padding=20)
        self.buttonSaveFrame.grid(row=0, column=0, sticky="nsew")
        self.buttonSaveFrame.grid_columnconfigure(0, weight=1)
        self.buttonSaveFrame.grid_columnconfigure(1, weight=5)
        self.buttonSaveFrame.grid_rowconfigure(0, weight=1)
        self.buttonSaveFrame.grid_rowconfigure(1, weight=1)

        # Ripristina questa riga per evitare errori di variabile non definita
        bg_color = self.style.lookup("MyCustomFrame.TFrame", "background")

        # Add graph
        self.fig, self.ax = plt.subplots()
        # self.ax.set_title("Prova a Trazione")
        self.ax.set_xlabel("Spostamento [mm]")
        self.ax.set_ylabel("Pressione [MPa]")

        # Creazione del canvas
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.graphFrame)
        canvas_widget = self.canvas.get_tk_widget()

        # Imposta il colore di sfondo
        canvas_widget.configure(bg=bg_color)
        self.fig.patch.set_facecolor(bg_color)

        # Posiziona il grafico
        canvas_widget.grid(row=1, column=0, sticky='NSEW')

    def resize_logo(self, event):
        """Dynamically resizes the logo based on window width"""
        self.new_height = int(root.winfo_height() / 20)  # Adaptive width
        self.new_width = int(self.new_height / self.aspect_ratio)

        resized_image = self.original_image.resize((self.new_width, self.new_height), Image.LANCZOS)
        self.logo_tk = ImageTk.PhotoImage(resized_image)

        self.logo_label.config(image=self.logo_tk)
        self.logo_label.image = self.logo_tk  # Keep reference

    def refresh_com_ports(self):
        """Update the dropdown with the latest available COM ports."""
        ports = serial.tools.list_ports.comports()
        self.com_ports = [port.device for port in ports]

        # Update dropdown values dynamically
        self.com_ports_dropdown["values"] = self.com_ports

        root.after(1000, self.refresh_com_ports)

    def connect_serial(self, event=None):
        com_port = self.com_port_var.get()
        if com_port:
            try:
                self.serial_obj = serial.Serial(com_port, 9600, timeout=1)
                print(f"connect_serial: connessione seriale avvenuta con successo!")
                self.serial_obj.reset_input_buffer()
                self.serial_obj.reset_output_buffer()
                if self.serial_obj and self.serial_obj.is_open:
                    self.helpLblVar.set("Comunicazione seriale attiva: lettura in corso...")
                    self.read_data()
                else:
                    self.helpLblVar.set("Errore: porta COM selezionata non corretta.")
                    root.after(300, self.connect_serial)
            except serial.SerialException:
                self.serial_obj = None
                print(f"connect_serial: errore durante la connessione seriale!")
                self.helpLblVar.set("Errore di connessione! Controlla la porta COM e il cavo USB.")
                root.after(300, self.connect_serial)
        else:
            print(f"conncetr_serial: nessuna porta COM trovata!")
            root.after(300, self.connect_serial)

    def read_data(self):
        """Read data from the serial port every 100ms."""
        
        """Legge i dati dalla porta seriale solo se la connessione è valida."""
        if not self.serial_obj or not self.serial_obj.is_open:
            # Ferma la lettura se la connessione è persa o se il ciclo è stato interrotto
            return

        if self.serial_obj.in_waiting > 0:
            try:
                # Try to read position and force data
                # position = self.serial_obj.readline().decode('utf-8').strip()
                # spring_position = self.serial_obj.readline().decode('utf-8').strip()

                # Leggi e valida due righe consecutive dalla seriale
                while True:
                    line1 = self.serial_obj.readline().decode('utf-8').strip()
                    if not line1.startswith("S1:"):
                        print(f"Linea inattesa (atteso S1): {line1}")
                        continue  # Scarta e riprova

                    line2 = self.serial_obj.readline().decode('utf-8').strip()
                    if not line2.startswith("S2:"):
                        print(f"Linea inattesa (atteso S2): {line2}")
                        continue  # Scarta e riprova

                    # Estrai solo il numero dopo S1: e S2:
                    try:
                        position = float(line1.replace("S1:", "").strip())
                        spring_position = float(line2.replace("S2:", "").strip())
                    except ValueError:
                        print(f"Errore di conversione: '{line1}' '{line2}'")
                        continue  # Scarta e riprova

                    break  # Esci dal ciclo dopo aver letto una coppia valida

                # Parse the data and apply zero correction
                pos = round(float(position) - self.zeropos, 2)  # Arrotonda a 2 cifre

                # RIMUOVI logica gainVar e radio button, usa direttamente spring_position come forza
                evalLoad = round(float(spring_position), 2)
            
                # Update GUI variables
                self.pos.set(pos)
                self.load.set(evalLoad)

                # Store the data in arrays for plotting
                if not self.stop:
                    self.arraypos.append(pos)
                    self.arrayforce.append(evalLoad)

                if not self.stop:
                    # Update the plot with new data
                    self.ax.clear()  # Clear the previous plot
                    self.ax.plot(self.arraypos, self.arrayforce)
                    self.ax.set_xlabel("Spostamento [mm]")
                    self.ax.set_ylabel("Forza [Kg]")
                    self.ax.set_title("Forza-Spostamento")
                    self.canvas.draw()

            except ValueError:
                # Handle any parsing errors (non-numeric data)
                print("read_data: (EXCEPT ERROR) Errore durante la lettura dei dati")

        # Schedule the next data reading
        root.after(50, self.read_data)  # Read data again in 50ms(20Hz)

    def start_acquisition(self):
        """Handler per il bottone Start: tara e avvia la registrazione."""
        if self.serial_obj and self.serial_obj.is_open:
            try:
                # Svuota il buffer per evitare dati "vecchi"
                self.serial_obj.reset_input_buffer()
                # Leggi due righe e assicurati che la prima sia S1
                while True:
                    line1 = self.serial_obj.readline().decode('utf-8').strip()
                    if not line1.startswith("S1:"):
                        print(f"Linea inattesa (atteso S1 per tara): {line1}")
                        continue
                    posizione_attuale = float(line1.replace("S1:", "").strip())
                    break
                # Scarta la riga S2 (non serve per la tara)
                _ = self.serial_obj.readline()
                self.zeropos = posizione_attuale
                # Svuota gli array dei dati per evitare dati pre-tara
                self.arraypos.clear()
                self.arrayforce.clear()
                # Pulisci il grafico
                self.ax.clear()
                self.ax.set_xlabel("Spostamento [mm]")
                self.ax.set_ylabel("Forza [Kg]")
                self.ax.set_title("Forza-Spostamento")
                self.canvas.draw()
                self.show_user_message("Acquisizione avviata. Tara effettuata.", timeout=3000)
                self.stop = False  # Sblocca la registrazione
            except Exception as e:
                self.show_user_message(f"Errore durante la tara: {e}", timeout=3000)
                self.zeropos = 0.0
                self.stop = True
        else:
            self.show_user_message("Seriale non connessa, impossibile registrare dati.", timeout=3000)
            self.stop = True
            return

    def stop_acquisition(self):
        """Handler per il bottone Stop: ferma la registrazione."""
        self.stop = True
        if self.arraypos and self.arrayforce:
            self.show_user_message("Acquisizione fermata. Puoi salvare il report.", timeout=3000)
        else:
            self.show_user_message("Acquisizione fermata. Nessun dato registrato.", timeout=3000)
    
    # --- RIMOSSO: Metodo save_report e riferimenti a reportName ---
    # def save_report(self):
    #     report_name = self.reportName.get().strip()
    #     if not report_name:
    #         messagebox.showerror("Errore", "Inserire un nome per il report.")
    #         return
    #     if not self.stop:
    #         messagebox.showerror("Errore", "Fermare l'acquisizione prima di salvare il report.")
    #         return
    #     if not self.arraypos or not self.arrayforce:
    #         messagebox.showerror("Errore", "Nessun dato da salvare.")
    #         return

    #     # Chiedi all'utente dove e come chiamare il file
    #     filename = filedialog.asksaveasfilename(
    #         title="Salva report come...",
    #         defaultextension=".xlsx",
    #         filetypes=[("Excel files", "*.xlsx")],
    #         initialfile=report_name
    #     )
    #     if not filename:
    #         return

    #     # Crea DataFrame
    #     df = pd.DataFrame({
    #         "Posizione [mm]": self.arraypos,
    #         "Forza [Kg]": self.arrayforce
    #     })

    #     # Trova il picco di forza e lo spostamento relativo
    #     max_force = max(self.arrayforce)
    #     idx_max = self.arrayforce.index(max_force)
    #     pos_at_max = self.arraypos[idx_max]

    #     # Salva temporaneamente il DataFrame
    #     df.to_excel(filename, index=False, startrow=4)  # Lascia spazio per info in alto

    #     # Ora modifica il file con openpyxl per aggiungere info e grafico
    #     wb = openpyxl.load_workbook(filename)
    #     ws = wb.active

    #     # Scrivi le info in alto
    #     ws["A1"] = "Picco di Forza [Kg]"
    #     ws["B1"] = max_force
    #     ws["A2"] = "Spostamento al Picco [mm]"
    #     ws["B2"] = pos_at_max

    #     # Crea il grafico a dispersione (scatter)
    #     chart = ScatterChart()
    #     chart.title = "Forza vs Spostamento"
    #     chart.x_axis.title = "Spostamento [mm]"
    #     chart.y_axis.title = "Forza [Kg]"

    #     # Imposta gli assi per incrociarsi a zero (se i dati lo permettono)
    #     chart.x_axis.crosses = "autoZero"
    #     chart.y_axis.crosses = "autoZero"
    #     chart.x_axis.crossAx = chart.y_axis.axId
    #     chart.y_axis.crossAx = chart.x_axis.axId

    #     # Riferimenti ai dati (attenzione: +6 perché i dati partono da riga 6)
    #     xvalues = Reference(ws, min_col=1, min_row=6, max_row=6+len(self.arraypos)-1)
    #     yvalues = Reference(ws, min_col=2, min_row=6, max_row=6+len(self.arrayforce)-1)
    #     series = Series(yvalues, xvalues)
    #     series.marker = None      # Nessun punto visibile
    #     series.smooth = True      # Linea curva smussata
    #     chart.series.append(series)

    #     # Rimuovi la legenda (opzionale, se non vuoi nemmeno la legenda)
    #     chart.legend = None
    #     ws.add_chart(chart, "D5")  # Posiziona il grafico

    #     wb.save(filename)

    #     messagebox.showinfo("Successo", f"Report salvato come:\n{filename}")
    
    def reset_help_message_if_disconnected(self):
        # Se la seriale non è connessa, mostra il messaggio di default
        if not self.serial_obj or not self.serial_obj.is_open:
            self.helpLblVar.set(self.default_help_message)
            # RIMUOVI: self.helpLblVar.set(self.gainVar)

    def show_user_message(self, message, timeout=3000):
        """Mostra un messaggio temporaneo nella label e poi ripristina lo stato seriale."""
        self.helpLblVar.set(message)
        # Cancella eventuali timer precedenti
        if hasattr(self, "_msg_timer") and self._msg_timer:
            self.root.after_cancel(self._msg_timer)
        # Dopo timeout ms, mostra lo stato seriale
        self._msg_timer = self.root.after(timeout, self.update_serial_status_message)

    def update_serial_status_message(self):
        """Mostra nella label lo stato attuale della connessione seriale."""
        if self.serial_obj and self.serial_obj.is_open:
            self.helpLblVar.set("Comunicazione seriale attiva: lettura in corso...")
        else:
            self.helpLblVar.set("Seriale non connessa. Collegare USB e selezionare porta COM.")

if __name__ == "__main__":
    root = ttkb.Window(themename="simplex")
    app = App(root)
    root.mainloop()